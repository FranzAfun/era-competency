import re

from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import user_passes_test
from django.db.models import Avg, Count, ExpressionWrapper, F, FloatField, Max, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from openpyxl import Workbook, load_workbook

from .models import Assessment, AssessmentCycle, Executive, Option, Question, Stage


def _is_portal_admin(user):
    return user.is_authenticated and (user.is_staff or user.is_superuser)


def _get_current_cycle():
    current_cycle = AssessmentCycle.objects.filter(is_current=True).order_by('-sequence').first()
    if current_cycle:
        return current_cycle

    latest_cycle = AssessmentCycle.objects.order_by('-sequence').first()
    next_sequence = latest_cycle.sequence + 1 if latest_cycle else 1
    return AssessmentCycle.objects.create(
        name=f'Assessment {next_sequence}',
        sequence=next_sequence,
        is_current=True,
    )


def admin_login_view(request):
    if _is_portal_admin(request.user):
        return redirect('admin_portal_dashboard')

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')

        if not username or not password:
            messages.error(request, 'Username and password are required.')
            return render(request, 'admin_portal/login.html')

        user = authenticate(request, username=username, password=password)
        if user and (user.is_staff or user.is_superuser):
            login(request, user)
            return redirect('admin_portal_dashboard')

        messages.error(request, 'Invalid admin credentials.')

    return render(request, 'admin_portal/login.html')


@user_passes_test(_is_portal_admin, login_url='admin_portal_login')
def admin_logout_view(request):
    logout(request)
    return redirect('admin_portal_login')


@user_passes_test(_is_portal_admin, login_url='admin_portal_login')
def admin_dashboard_view(request):
    current_cycle = _get_current_cycle()
    stage_stats = []
    for stage in Stage.objects.order_by('order'):
        question_count = Question.objects.filter(stage_ref=stage).count()
        assessments = Assessment.objects.filter(stage_ref=stage)
        total_attempts = assessments.count()
        pass_count = assessments.filter(passed=True).count()
        average_score = assessments.aggregate(avg=Avg('score'))['avg'] or 0
        pass_rate = round((pass_count / total_attempts) * 100, 2) if total_attempts else 0

        stage_stats.append({
            'stage': stage,
            'question_count': question_count,
            'total_attempts': total_attempts,
            'pass_rate': pass_rate,
            'average_score': round(average_score, 2),
        })

    context = {
        'total_executives': Executive.objects.count(),
        'total_questions': Question.objects.count(),
        'total_assessments': Assessment.objects.count(),
        'overall_pass_rate': _overall_pass_rate(),
        'current_cycle': current_cycle,
        'stage_stats': stage_stats,
        'recent_assessments': Assessment.objects.select_related('executive', 'stage_ref', 'cycle').order_by('-created_at')[:10],
        'difficult_questions': _get_difficult_questions(),
    }

    return render(request, 'admin_portal/dashboard.html', context)


@user_passes_test(_is_portal_admin, login_url='admin_portal_login')
def admin_stages_view(request):
    if request.method == 'POST':
        edit_stage_id_raw = request.POST.get('edit_stage_id', '').strip()
        name = request.POST.get('name', '').strip()
        order_raw = request.POST.get('order', '').strip()
        is_active = request.POST.get('is_active') == 'on'

        if not name or not order_raw:
            messages.error(request, 'Stage name and order are required.')
            return redirect('admin_portal_stages')

        try:
            order = int(order_raw)
        except ValueError:
            messages.error(request, 'Stage order must be a number.')
            return redirect('admin_portal_stages')

        if order < 1 or order > 4:
            messages.error(request, 'Stage order must be between 1 and 4.')
            return redirect('admin_portal_stages')

        if edit_stage_id_raw:
            try:
                edit_stage_id = int(edit_stage_id_raw)
            except ValueError:
                messages.error(request, 'Invalid stage selected for update.')
                return redirect('admin_portal_stages')

            stage = Stage.objects.filter(id=edit_stage_id).first()
            if stage is None:
                messages.error(request, 'Stage not found for update.')
                return redirect('admin_portal_stages')

            if Stage.objects.exclude(id=stage.id).filter(order=order).exists():
                messages.error(request, f'Stage order {order} is already in use. Edit that stage directly or choose another order.')
                return redirect('admin_portal_stages')

            if stage.order != order and stage.assessments.exists():
                messages.error(request, f'{stage} already has assessment history. Its order cannot be changed; update the name or active status instead.')
                return redirect('admin_portal_stages')

            previous_label = str(stage)
            previous_order = stage.order
            stage.name = name
            stage.order = order
            stage.is_active = is_active
            stage.save(update_fields=['name', 'order', 'is_active'])

            if previous_order != order:
                Question.objects.filter(stage_ref=stage).update(stage=order)
                Assessment.objects.filter(stage_ref=stage).update(stage=order)

            messages.success(request, f'{previous_label} updated successfully.')
            return redirect('admin_portal_stages')

        if Stage.objects.filter(order=order).exists():
            messages.error(request, f'Stage order {order} is already in use. Edit the existing stage instead of creating a replacement.')
            return redirect('admin_portal_stages')

        stage = Stage.objects.create(name=name, order=order, is_active=is_active)
        messages.success(request, f'{stage} created successfully.')

        return redirect('admin_portal_stages')

    stages = Stage.objects.annotate(
        question_count=Count('questions', distinct=True),
        assessment_count=Count('assessments', distinct=True),
    ).order_by('order')
    return render(request, 'admin_portal/stages.html', {'stages': stages})


@user_passes_test(_is_portal_admin, login_url='admin_portal_login')
def admin_reset_cycle_view(request):
    if request.method != 'POST':
        return redirect('admin_portal_stages')

    current_cycle = _get_current_cycle()
    latest_cycle = AssessmentCycle.objects.order_by('-sequence').first()
    next_sequence = latest_cycle.sequence + 1 if latest_cycle else 1

    AssessmentCycle.objects.filter(id=current_cycle.id).update(is_current=False)
    new_cycle = AssessmentCycle.objects.create(
        name=f'Assessment {next_sequence}',
        sequence=next_sequence,
        is_current=True,
    )

    messages.success(request, f'Started {new_cycle.name}. Historical assessment records were preserved.')
    return redirect('admin_portal_stages')


@user_passes_test(_is_portal_admin, login_url='admin_portal_login')
def admin_clear_stage_questions_view(request, stage_id):
    if request.method != 'POST':
        return redirect('admin_portal_stages')

    stage = get_object_or_404(Stage, id=stage_id)
    questions = Question.objects.filter(stage_ref=stage)
    question_count = questions.count()

    if question_count == 0:
        messages.error(request, f'{stage} does not have any questions to clear.')
    else:
        questions.delete()
        messages.success(
            request,
            f'Cleared {question_count} question(s) from {stage}. You can now rebuild the stage question set from scratch.',
        )

    return redirect('admin_portal_stages')


@user_passes_test(_is_portal_admin, login_url='admin_portal_login')
def admin_delete_stage_view(request, stage_id):
    if request.method != 'POST':
        return redirect('admin_portal_stages')

    stage = get_object_or_404(Stage, id=stage_id)
    if stage.questions.exists():
        messages.error(
            request,
            f'{stage} still has questions. Clear its questions first before deleting the stage.',
        )
        return redirect('admin_portal_stages')

    if stage.assessments.exists():
        messages.error(
            request,
            f'{stage} has assessment history and cannot be deleted. Set it inactive instead if you need to retire it.',
        )
        return redirect('admin_portal_stages')

    stage_label = str(stage)
    stage.delete()
    messages.success(request, f'Deleted {stage_label}.')

    return redirect('admin_portal_stages')


@user_passes_test(_is_portal_admin, login_url='admin_portal_login')
def admin_questions_view(request):
    active_stages = Stage.objects.filter(is_active=True).order_by('order')

    if request.method == 'POST':
        if request.POST.get('bulk_delete_mode') == '1':
            selected_ids_raw = request.POST.getlist('selected_question_ids')
            if not selected_ids_raw:
                messages.error(request, 'Select at least one question to delete.')
                return redirect('admin_portal_questions')

            selected_ids = []
            for raw_id in selected_ids_raw:
                try:
                    selected_ids.append(int(raw_id))
                except ValueError:
                    continue

            if not selected_ids:
                messages.error(request, 'Selected question IDs are invalid.')
                return redirect('admin_portal_questions')

            questions = Question.objects.filter(id__in=selected_ids).select_related('stage_ref')
            question_count = questions.count()

            if question_count == 0:
                messages.error(request, 'No matching questions were found to delete.')
                return redirect('admin_portal_questions')

            affected_stage_ids = set(
                questions.exclude(stage_ref__isnull=True).values_list('stage_ref_id', flat=True)
            )

            questions.delete()

            for stage_id in affected_stage_ids:
                remaining = Question.objects.filter(stage_ref_id=stage_id).order_by('order', 'id')
                for i, q in enumerate(remaining, start=1):
                    if q.order != i:
                        q.order = i
                        q.save(update_fields=['order'])

            messages.success(request, f'Deleted {question_count} question(s).')
            return redirect('admin_portal_questions')

        if request.POST.get('upload_mode') == '1':
            stage_id = request.POST.get('upload_stage_id', '').strip()
            upload_file = request.FILES.get('questions_file')

            if not stage_id or upload_file is None:
                messages.error(request, 'Please select a stage and upload an Excel file.')
                return redirect('admin_portal_questions')

            try:
                stage = Stage.objects.get(id=stage_id, is_active=True)
            except Stage.DoesNotExist:
                messages.error(request, 'Selected stage is invalid or inactive.')
                return redirect('admin_portal_questions')

            filename = upload_file.name.lower()
            if not filename.endswith('.xlsx'):
                messages.error(request, 'Only .xlsx files are supported for bulk upload.')
                return redirect('admin_portal_questions')

            try:
                imported_count, skipped_count, remaining_slots = _import_questions_from_xlsx(upload_file, stage)
            except ValueError as exc:
                messages.error(request, str(exc))
                return redirect('admin_portal_questions')

            if imported_count:
                messages.success(request, f'Imported {imported_count} questions into {stage}.')
            if skipped_count:
                messages.error(request, f'Skipped {skipped_count} row(s) due to invalid or incomplete data.')
            if remaining_slots == 0:
                messages.success(request, f'{stage} is now full at 25 questions.')

            return redirect('admin_portal_questions')

        stage_id = request.POST.get('stage_id', '').strip()
        edit_question_id_raw = request.POST.get('edit_question_id', '').strip()
        text = request.POST.get('text', '').strip()
        explanation = request.POST.get('explanation', '').strip()
        correct_option_raw = request.POST.get('correct_option', '').strip()

        options = [
            request.POST.get('option_1', '').strip(),
            request.POST.get('option_2', '').strip(),
            request.POST.get('option_3', '').strip(),
            request.POST.get('option_4', '').strip(),
        ]

        if not stage_id or not text or not explanation or not correct_option_raw or any(not opt for opt in options):
            messages.error(request, 'Please complete all question fields, including explanation and options.')
            return redirect('admin_portal_questions')

        try:
            stage = Stage.objects.get(id=stage_id, is_active=True)
            correct_option = int(correct_option_raw)
        except (Stage.DoesNotExist, ValueError):
            messages.error(request, 'Invalid stage or correct option selection.')
            return redirect('admin_portal_questions')

        if correct_option < 1 or correct_option > 4:
            messages.error(request, 'Correct option must be between 1 and 4.')
            return redirect('admin_portal_questions')

        if edit_question_id_raw:
            try:
                edit_question_id = int(edit_question_id_raw)
            except ValueError:
                messages.error(request, 'Invalid question selected for update.')
                return redirect('admin_portal_questions')

            question = Question.objects.filter(id=edit_question_id).select_related('stage_ref').first()
            if question is None:
                messages.error(request, 'Question not found for update.')
                return redirect('admin_portal_questions')

            existing_for_stage = Question.objects.filter(stage_ref=stage).exclude(id=question.id).count()
            if existing_for_stage >= 25:
                messages.error(request, f'{stage} already has 25 questions.')
                return redirect('admin_portal_questions')

            previous_stage_id = question.stage_ref_id
            previous_order = question.order

            question.text = text
            question.explanation = explanation
            if question.stage_ref_id != stage.id:
                next_order = (Question.objects.filter(stage_ref=stage).aggregate(max_order=Max('order'))['max_order'] or 0) + 1
                question.stage_ref = stage
                question.stage = stage.order
                question.order = next_order
                question.save(update_fields=['text', 'explanation', 'stage_ref', 'stage', 'order'])
                _reorder_stage_questions(previous_stage_id)
            else:
                question.stage = stage.order
                question.save(update_fields=['text', 'explanation', 'stage'])

            existing_options = list(question.options.order_by('id'))
            if len(existing_options) != 4:
                question.options.all().delete()
                Option.objects.bulk_create([
                    Option(question=question, text=options[0], is_correct=correct_option == 1),
                    Option(question=question, text=options[1], is_correct=correct_option == 2),
                    Option(question=question, text=options[2], is_correct=correct_option == 3),
                    Option(question=question, text=options[3], is_correct=correct_option == 4),
                ])
            else:
                for idx, option_obj in enumerate(existing_options, start=1):
                    option_obj.text = options[idx - 1]
                    option_obj.is_correct = correct_option == idx
                Option.objects.bulk_update(existing_options, ['text', 'is_correct'])

            current_order = question.order
            current_stage_label = question.stage_ref.name if question.stage_ref else f'Stage {question.stage}'
            if previous_stage_id == question.stage_ref_id:
                messages.success(request, f'Updated Question #{current_order} in {current_stage_label}.')
            else:
                messages.success(
                    request,
                    f'Updated Question #{previous_order} and moved it to Question #{current_order} in {current_stage_label}.',
                )
            return redirect('admin_portal_questions')

        existing_for_stage = Question.objects.filter(stage_ref=stage).count()
        if existing_for_stage >= 25:
            messages.error(request, f'{stage} already has 25 questions.')
            return redirect('admin_portal_questions')

        next_order = (Question.objects.filter(stage_ref=stage).aggregate(max_order=Max('order'))['max_order'] or 0) + 1

        question = Question.objects.create(
            text=text,
            explanation=explanation,
            stage=stage.order,
            stage_ref=stage,
            order=next_order,
        )

        Option.objects.bulk_create([
            Option(question=question, text=options[0], is_correct=correct_option == 1),
            Option(question=question, text=options[1], is_correct=correct_option == 2),
            Option(question=question, text=options[2], is_correct=correct_option == 3),
            Option(question=question, text=options[3], is_correct=correct_option == 4),
        ])

        messages.success(request, f'Question #{next_order} added to {stage}.')
        return redirect('admin_portal_questions')

    question_rows = list(Question.objects.select_related('stage_ref').prefetch_related('options').order_by('-id')[:100])
    for question in question_rows:
        ordered_options = list(question.options.all().order_by('id'))
        while len(ordered_options) < 4:
            ordered_options.append(None)

        question.option_text_1 = ordered_options[0].text if ordered_options[0] else ''
        question.option_text_2 = ordered_options[1].text if ordered_options[1] else ''
        question.option_text_3 = ordered_options[2].text if ordered_options[2] else ''
        question.option_text_4 = ordered_options[3].text if ordered_options[3] else ''
        question.explanation_text = question.explanation or ''
        question.correct_option_index = ''

        for idx, option_obj in enumerate(ordered_options[:4], start=1):
            if option_obj and option_obj.is_correct:
                question.correct_option_index = idx
                break

        question.stage_label = question.stage_ref.name if question.stage_ref else f'Stage {question.stage}'

    return render(request, 'admin_portal/questions.html', {
        'stages': active_stages,
        'question_rows': question_rows,
    })


@user_passes_test(_is_portal_admin, login_url='admin_portal_login')
def admin_delete_question_view(request, question_id):
    if request.method == 'POST':
        question = get_object_or_404(Question, id=question_id)
        stage = question.stage_ref
        question.delete()

        _reorder_stage_questions(stage.id if stage else None)

        messages.success(request, 'Question deleted successfully.')

    return redirect('admin_portal_questions')


@user_passes_test(_is_portal_admin, login_url='admin_portal_login')
def admin_questions_template_download_view(request):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Questions Template'
    sheet.append(['question_text', 'explanation', 'option_1', 'option_2', 'option_3', 'option_4', 'correct_option'])
    sheet.append(['What is 2 + 2?', '2 + 2 equals 4 because basic addition combines two pairs.', '3', '4', '5', '6', 2])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="questions_template.xlsx"'
    workbook.save(response)
    return response


def _overall_pass_rate():
    total = Assessment.objects.count()
    if total == 0:
        return 0
    passed = Assessment.objects.filter(passed=True).count()
    return round((passed / total) * 100, 2)


def _get_difficult_questions():
    return (
        Question.objects.annotate(
            total_answers=Count('response'),
            wrong_answers=Count('response', filter=Q(response__is_correct=False)),
        )
        .filter(total_answers__gte=3)
        .annotate(
            wrong_rate=ExpressionWrapper(
                100.0 * F('wrong_answers') / F('total_answers'),
                output_field=FloatField(),
            )
        )
        .select_related('stage_ref')
        .order_by('-wrong_rate', '-wrong_answers', 'id')[:10]
    )


def _import_questions_from_xlsx(upload_file, stage):
    workbook = load_workbook(upload_file, read_only=True, data_only=True)
    worksheet = workbook.active

    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError('The uploaded file is empty.')

    header = [_normalize_header(cell) for cell in rows[0]]
    required_header = ['question_text', 'explanation', 'option_1', 'option_2', 'option_3', 'option_4', 'correct_option']
    if header[:7] != required_header:
        raise ValueError(
            'Invalid template columns. Required columns are: question_text, explanation, option_1, option_2, option_3, option_4, correct_option.'
        )

    existing_count = Question.objects.filter(stage_ref=stage).count()
    remaining_slots = max(0, 25 - existing_count)
    if remaining_slots == 0:
        raise ValueError(f'{stage} already has 25 questions.')

    next_order = (Question.objects.filter(stage_ref=stage).aggregate(max_order=Max('order'))['max_order'] or 0) + 1
    imported_count = 0
    skipped_count = 0

    for row in rows[1:]:
        if imported_count >= remaining_slots:
            break

        normalized = [str(value).strip() if value is not None else '' for value in row[:7]]
        if not any(normalized):
            continue

        if len(normalized) < 7 or any(not item for item in normalized[:6]):
            skipped_count += 1
            continue

        question_text, explanation, option_1, option_2, option_3, option_4, correct_option_raw = normalized[:7]
        if not question_text:
            skipped_count += 1
            continue

        try:
            correct_option = int(correct_option_raw)
        except ValueError:
            skipped_count += 1
            continue

        if correct_option not in (1, 2, 3, 4):
            skipped_count += 1
            continue

        question = Question.objects.create(
            text=question_text,
            explanation=explanation,
            stage=stage.order,
            stage_ref=stage,
            order=next_order,
        )

        Option.objects.bulk_create([
            Option(question=question, text=option_1, is_correct=correct_option == 1),
            Option(question=question, text=option_2, is_correct=correct_option == 2),
            Option(question=question, text=option_3, is_correct=correct_option == 3),
            Option(question=question, text=option_4, is_correct=correct_option == 4),
        ])

        imported_count += 1
        next_order += 1

    remaining_after_import = max(0, 25 - (existing_count + imported_count))
    return imported_count, skipped_count, remaining_after_import


def _normalize_header(value):
    raw = str(value or '').strip().lower()
    raw = re.sub(r'\s+', '_', raw)
    return raw


def _reorder_stage_questions(stage_id):
    if not stage_id:
        return

    remaining = Question.objects.filter(stage_ref_id=stage_id).order_by('order', 'id')
    for index, question in enumerate(remaining, start=1):
        if question.order != index:
            question.order = index
            question.save(update_fields=['order'])
